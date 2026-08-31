from pathlib import Path

from PySide6.QtCore import Qt

from pulse import app
from pulse.extensions import SUPPORTED_SPREADSHEET_EXTENSIONS
from pulse.interface.ui_generated.model.setup.fluid.load_fluid_composition_ui import (
    LoadFluidComposition_UI,
)
from pulse.interface.user_input.data_handler.file_dialog_service import (
    FileDialogService,
)
from pulse.interface.user_input.project.print_message import PrintMessageInput


class LoadFluidCompositionInput(LoadFluidComposition_UI):
    def __init__(self, file_path: str = ""):
        super().__init__()

        app().main_window.set_input_widget(self)

        self.file_path = file_path
       
        self._initialize()
        self._config_window()
        self._create_connections()

        self._config_widgets()
        self._load_file()
        self.exec()

    def _initialize(self):

        self.complete = False
        self.imported_data = {}
        self.fluid_composition_data: list[tuple[int, str, str, str]] = []
        self.state_properties_data: list[tuple[int, str, str, str]] = []

        self.desktop_path = Path.home() / "Desktop"

    def _config_window(self):
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)
        self.setWindowIcon(app().main_window.pulse_icon)
        self.setWindowTitle("OpenPulse")

    def _create_connections(self):
        self.pushButton_cancel.clicked.connect(self.close)
        self.pushButton_load_composition.clicked.connect(self.confirm_button_callback)
        self.pushButton_search.clicked.connect(self.search_button_callback)

    def _config_widgets(self):
        self.lineEdit_file_path.setDisabled(True)
        self.comboBox_sheet_names.setDisabled(True)
        self.comboBox_state_properties.setDisabled(True)

    def _load_file(self):
        if Path(self.file_path).exists():
            self.lineEdit_file_path.setText(self.file_path)
            self.load_composition_data_from_file()

    def search_button_callback(self):
        caption = "Open the fluid composition file"

        file_path = FileDialogService.open_file(SUPPORTED_SPREADSHEET_EXTENSIONS, caption, "fluid_composition_folder")

        if file_path is None:
            self.file_path = ""
            return
        
        if isinstance(file_path, Path):
            self.file_path = str(file_path)

        self.lineEdit_file_path.setText(self.file_path)
        self.load_composition_data_from_file()

    def load_composition_data_from_file(self):

        if self.lineEdit_file_path.text() == "":
            return

        self.imported_data.clear()
        self.comboBox_sheet_names.clear()
        self.comboBox_state_properties.clear()

        from openpyxl import load_workbook
        from polars import read_excel

        wb = load_workbook(self.file_path)

        for sheetname in wb.sheetnames:

            try:
                sheet_data = read_excel(
                    self.file_path,
                    sheet_name=sheetname,
                    columns=(0, 1, 2, 3),
                    has_header=True,
                )

                if "state properties" in sheetname.lower().replace("_", " "):
                    self.comboBox_state_properties.addItem(sheetname)
                    if not self.comboBox_state_properties.isEnabled():
                        self.comboBox_state_properties.setDisabled(False)
                    
                else:
                    self.comboBox_sheet_names.addItem(sheetname)
                    if not self.comboBox_sheet_names.isEnabled():
                        self.comboBox_sheet_names.setDisabled(False)

                self.imported_data[sheetname] = sheet_data.to_numpy()

            except Exception as error_log:
                window_title = "Error"
                title = "Error while reading data from file"
                message = str(error_log)
                PrintMessageInput([window_title, title, message])
                return True

    def confirm_button_callback(self):
        if not self.imported_data:
            return

        composition_key = self.comboBox_sheet_names.currentText()
        state_properties_key = self.comboBox_state_properties.currentText()

        self.fluid_composition_data = self.imported_data.get(composition_key)
        self.state_properties_data = self.imported_data.get(state_properties_key)

        self.complete = True
        self.close()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.confirm_button_callback()
            return

        if event.key() == Qt.Key_Escape:
            self.close()