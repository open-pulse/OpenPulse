# fmt: off

from PyQt5.QtWidgets import QComboBox, QDialog, QLabel, QLineEdit, QPushButton, QTabWidget, QTreeWidget, QTreeWidgetItem
from PyQt5.QtGui import QCloseEvent
from PyQt5.QtCore import Qt, QEvent, QObject, pyqtSignal
from PyQt5 import uic

from pulse import app, UI_DIR
from pulse.interface.user_input.project.print_message import PrintMessageInput
from pulse.interface.user_input.project.get_user_confirmation_input import GetUserConfirmationInput

import os
import numpy as np
from pathlib import Path

window_title_1 = "Error"
window_title_2 = "Warning"


class AddAcousticTransferElementInput(QDialog):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        ui_path = UI_DIR / "model/setup/acoustic/acoustic_transfer_element_input.ui"
        uic.loadUi(ui_path, self)

        app().main_window.set_input_widget(self)
        self.properties = app().project.model.properties
        self.preprocessor = app().project.model.preprocessor

        self._config_window()
        self._initialize()
        self._define_qt_variables()
        self._create_connections()
        # self._config_widgets()
        self.load_nodal_info()
        self.selection_callback()

        while self.keep_window_open:
            self.exec()

    def _config_window(self):
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)
        self.setWindowIcon(app().main_window.pulse_icon)
        self.setWindowTitle("OpenPulse")

    def _initialize(self):

        self.keep_window_open = True

        self.type_label = None
        self.dkey = None
        self.log_removal = True

        self.element_transfer_data = dict()

        self.before_run = app().project.get_pre_solution_model_checks()
    
    def _define_qt_variables(self):

        # QComboBox
        self.comboBox_data_type:  QComboBox

        # QLabel
        self.label_selection: QLabel

        # QLineEdit
        self.lineEdit_input_node_id: QLineEdit
        self.lineEdit_output_node_id: QLineEdit
        self.lineEdit_selected_id: QLineEdit
        self.lineEdit_spreadsheet_path: QLineEdit
        self.current_lineEdit = self.lineEdit_output_node_id

        # QPushButton
        self.pushButton_attribute: QPushButton
        self.pushButton_cancel: QPushButton
        self.pushButton_invert_selection: QPushButton
        self.pushButton_remove: QPushButton
        self.pushButton_reset: QPushButton
        self.pushButton_search: QPushButton

        # QTabWidget
        self.tabWidget_main: QTabWidget

        # QTreeWidget
        self.treeWidget_nodal_info: QTreeWidget

    def _create_connections(self):
        #
        self.pushButton_attribute.clicked.connect(self.attribute_callback)
        self.pushButton_invert_selection.clicked.connect(self.invert_selection_callback)
        self.pushButton_cancel.clicked.connect(self.close)
        self.pushButton_remove.clicked.connect(self.remove_callback)
        self.pushButton_reset.clicked.connect(self.reset_callback)
        self.pushButton_search.clicked.connect(self.search_callback)
        #
        self.tabWidget_main.currentChanged.connect(self.tab_event_callback)
        #
        self.treeWidget_nodal_info.itemClicked.connect(self.on_click_item)
        self.treeWidget_nodal_info.itemDoubleClicked.connect(self.on_doubleclick_item)
        #
        self.clickable(self.lineEdit_input_node_id).connect(self.lineEdit_1_clicked)
        self.clickable(self.lineEdit_output_node_id).connect(self.lineEdit_2_clicked)
        #
        app().main_window.selection_changed.connect(self.selection_callback)

    def selection_callback(self):
        selected_nodes = selected_nodes = app().main_window.list_selected_nodes()
        if selected_nodes:
            if len(selected_nodes) == 1:
                node_id = selected_nodes[0]
                self.current_lineEdit.setText(str(node_id))                

    def clickable(self, widget):
        class Filter(QObject):
            clicked = pyqtSignal()

            def eventFilter(self, obj, event):
                if obj == widget and event.type() == QEvent.MouseButtonRelease and obj.rect().contains(event.pos()):
                    self.clicked.emit()
                    return True
                else:
                    return False

        filter = Filter(widget)
        widget.installEventFilter(filter)
        return filter.clicked

    def lineEdit_1_clicked(self):
        self.current_lineEdit = self.lineEdit_input_node_id

    def lineEdit_2_clicked(self):
        self.current_lineEdit = self.lineEdit_output_node_id

    def invert_selection_callback(self):
        temp_text_input = self.lineEdit_input_node_id.text()
        temp_text_output = self.lineEdit_output_node_id.text()
        self.lineEdit_input_node_id.setText(temp_text_output)
        self.lineEdit_output_node_id.setText(temp_text_input) 

    def attribute_callback(self):

        path = self.lineEdit_spreadsheet_path.text()

        if path == "":
            if self.search_callback():
                return
            
        if self.check_inputs():
            return

        if os.path.exists(path):
            self.import_element_transfer_data(path)

            if self.element_transfer_data:
                self.process_acoustic_element_transfer_data(path)
                self.actions_to_finalize()

    def remove_table_files_from_nodes(self, node_id : list):
        table_names = self.properties.get_nodal_related_table_names("acoustic_transfer_element", node_id)
        self.process_table_file_removal(table_names)

    def process_table_file_removal(self, table_names : list):
        if table_names:
            for table_name in table_names:
                self.properties.remove_imported_tables("acoustic", table_name)
            app().pulse_file.write_imported_table_data_in_file()

    def remove_callback(self):

        if  self.lineEdit_selected_id.text() != "":

            linked_nodes = self.lineEdit_selected_id.text()
            node_ids = [int(node_id) for node_id in linked_nodes.split("-")]

            self.remove_table_files_from_nodes(node_ids)
            self.properties._remove_nodal_property("acoustic_transfer_element", node_ids)

            self.actions_to_finalize()
            # self.close()

    def reset_callback(self):

            self.hide()

            title = f"Resetting of acoustic transfer element"
            message = "Would you like to remove all acoustic transfer element from the acoustic model?"

            buttons_config = {"left_button_label" : "No", "right_button_label" : "Yes"}
            read = GetUserConfirmationInput(title, message, buttons_config=buttons_config)

            if read._cancel:
                return

            if read._continue:

                for (property, *args) in self.properties.nodal_properties.keys():
                    if property == "acoustic_transfer_element":
                        self.remove_table_files_from_nodes(args)                    

                self.properties._reset_nodal_property("acoustic_transfer_element")

                self.actions_to_finalize()
                # self.close()

    def search_callback(self):

        last_path = app().main_window.config.get_last_folder_for("imported table folder")
        if last_path is None:
            last_path = str(Path().home())

        caption = f"Choose a file to import element transfer data"
        path, check = app().main_window.file_dialog.get_open_file_name(
                                                                        caption, 
                                                                        last_path, 
                                                                        'Table File (*.xls; *.xlsx;)'
                                                                        )

        if not check:
            return True

        self.lineEdit_spreadsheet_path.setText(path)
        app().main_window.config.write_last_folder_path_in_file("imported table folder", path)

    def check_inputs(self):

        input_node_id = self.lineEdit_input_node_id.text()
        stop, self.input_node_id = self.before_run.check_selected_ids(input_node_id, "nodes", single_id=True)
        if stop:
            self.lineEdit_input_node_id.setFocus()
            return True
        
        output_node_id = self.lineEdit_output_node_id.text()
        stop, self.output_node_id = self.before_run.check_selected_ids(output_node_id, "nodes", single_id=True)
        if stop:
            self.lineEdit_output_node_id.setFocus()
            return True

    def import_element_transfer_data(self, imported_path: str):

        from pandas import read_excel
        from openpyxl import load_workbook

        self.element_transfer_data.clear()

        try:

            sufix = Path(imported_path).suffix
            # filename = os.path.basename(imported_path)

            if sufix in [".xls", ".xlsx"]:
                wb = load_workbook(imported_path)
                sheetnames = wb.sheetnames

                if self.comboBox_data_type.currentIndex() == 0:
                    cols = list(np.arange(3))
                else:
                    cols = list(np.arange(9))

                for sheetname in sheetnames:

                    sheet_data = read_excel(
                                            imported_path, 
                                            sheet_name = sheetname, 
                                            header = 0, 
                                            usecols = cols
                                            ).to_numpy()

                    self.element_transfer_data[sheetname] = sheet_data

        except Exception as log_error:
            title = "Error while loading data from file"
            message = str(log_error)
            PrintMessageInput([window_title_1, title, message])
            return
        
    def update_frequency_setup(self, values: np.ndarray, path: str):

        self.frequencies = values[:, 0]
        f_min = self.frequencies[0]
        f_max = self.frequencies[-1]
        f_step = self.frequencies[1] - self.frequencies[0]

        if app().project.model.change_analysis_frequency_setup(list(self.frequencies)):

            self.lineEdit_spreadsheet_path.setText("")

            title = "Project frequency setup cannot be modified"
            message = f"The following imported table of values has a frequency setup\n"
            message += "different from the others already imported ones. The current\n"
            message += "project frequency setup is not going to be modified."
            message += f"\n\n{os.path.basename(path)}"
            PrintMessageInput([window_title_1, title, message])
            return None, None

        else:

            frequency_setup = { "f_min" : f_min,
                                "f_max" : f_max,
                                "f_step" : f_step }

            app().project.model.set_frequency_setup(frequency_setup)

    def process_acoustic_element_transfer_data(self, path):

        aux = dict()
        table_names = list()
        for k, (sheetaname, et_data) in enumerate(self.element_transfer_data.items()):

            if k == 0:
                self.update_frequency_setup(et_data, path)

            if self.comboBox_data_type.currentIndex() == 1:                 
                if et_data.shape[1] == 9:
                    for i, e_label in enumerate(["a11", "a12", "a21", "a22"]):
                        data_ij = np.array([et_data[:,0], et_data[:,2*i+1], et_data[:,2*i+2]], dtype=float).T
                        table_name = f"admittance_matrix_data_{e_label}_nodes_{self.input_node_id}_{self.output_node_id}"
                        aux[e_label] = {"values" : data_ij,
                                        "table_name" : table_name}
                else:
                    continue

            else:

                linked_nodes = f"{self.input_node_id}_{self.output_node_id}"

                if "H11" in sheetaname:
                    table_name = f"transfer_function_H11_nodes_{linked_nodes}"
                    aux["H11"] = {"values" : et_data,
                                  "table_name" : table_name}

                elif "H21" in sheetaname:
                    table_name = f"transfer_function_H21_nodes_{linked_nodes}"
                    aux["H21"] = {"values" : et_data,
                                  "table_name" : table_name}

                elif "H12" in sheetaname:
                    table_name = f"transfer_function_H12_nodes_{linked_nodes}"
                    aux["H12"] = {"values" : et_data,
                                  "table_name" : table_name}

                elif "H22" in sheetaname:
                    table_name = f"transfer_function_H22_nodes_{linked_nodes}"
                    aux["H22"] = {"values" : et_data,
                                  "table_name" : table_name}

                else:
                    continue

        for _data in aux.values():
            values = _data["values"]
            table_name = _data["table_name"]
            self.properties.add_imported_tables("acoustic", table_name, values)

        coords = list()
        node_ids = [self.input_node_id, self.output_node_id]
        for node_id in node_ids:
            node = app().project.model.preprocessor.nodes[node_id]
            coords.extend(list(np.round(node.coordinates, 5)))

        table_names = list()

        if self.comboBox_data_type.currentIndex() == 0:
            data_source = "transfer_functions"
            for key in ["H11", "H21", "H12", "H22"]:
                table_names.append(aux[key]["table_name"])

        else:
            data_source = "admittance_matrix"
            for key in ["a11", "a12", "a21", "a22"]:
                table_names.append(aux[key]["table_name"])

        data = {
                "coords" : coords,
                "table_names" : table_names,
                "table_paths" : [path],
                "element_transfer_data_source" : data_source
                }

        self.properties._set_nodal_property("acoustic_transfer_element", data, node_ids)

    def actions_to_finalize(self):
        app().pulse_file.write_nodal_properties_in_file()
        app().pulse_file.write_imported_table_data_in_file()
        self.load_nodal_info()
        app().main_window.update_plots(reset_camera=False)

    def on_click_item(self, item):
        input_node_id = int(item.text(1))
        output_node_id = int(item.text(2))
        self.pushButton_remove.setEnabled(True)
        self.lineEdit_selected_id.setText(f"{input_node_id}-{output_node_id}")
        app().main_window.set_selection(nodes=(input_node_id, output_node_id))

    def on_doubleclick_item(self, item):
        self.on_click_item(item)

    def tab_event_callback(self):
        self.lineEdit_selected_id.setText("")
        self.pushButton_remove.setDisabled(True)
        # if self.tabWidget_main.currentIndex() == 1:
        #     self.lineEdit_selected_id.setText("")
        # else:
        #     self.selection_callback()

    def load_nodal_info(self):

        index = 0
        self.treeWidget_nodal_info.clear()

        for (property, *args), data in self.properties.nodal_properties.items():
            if property == "acoustic_transfer_element":
                if "values" in data.keys():
                    index += 1
                    new = QTreeWidgetItem([str(index), str(args[0]), str(args[1])])
                    for i in range(3):
                        new.setTextAlignment(i, Qt.AlignCenter)
                    self.treeWidget_nodal_info.addTopLevelItem(new)

        self.tabWidget_main.setTabVisible(1, False)
        for (_property, *_) in self.properties.nodal_properties.keys():
            if _property == "acoustic_transfer_element":
                self.tabWidget_main.setCurrentIndex(0)
                self.tabWidget_main.setTabVisible(1, True)
                return

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.attribute_callback()
        elif event.key() == Qt.Key_Delete:
            self.remove_callback()
        elif event.key() == Qt.Key_Escape:
            self.close()

    def closeEvent(self, a0: QCloseEvent | None) -> None:
        self.keep_window_open = False
        return super().closeEvent(a0)

# fmt: on