from PySide6.QtWidgets import QFileDialog

from pulse.interface.user_input.data_handler.imported_file import ImportedFile

from typing import List
from pathlib import Path
from polars import DataFrame as PolarsDataFrame
import numpy as np
import platform
import os
import h5py


class FileManager:

    @staticmethod
    def import_single_file(file_extensions: List[str], caption: str = "Open file", last_folder: str = None) -> ImportedFile | None:
        imported_data = FileManager.__import_files(caption, last_folder, file_extensions)

        if isinstance(imported_data, list):
            if imported_data:
                return imported_data[0]
        return None
    
    @staticmethod
    def import_multiple_files(file_extensions: List[str], caption: str = "Open file", last_folder: str = None) -> List[ImportedFile]:
        return FileManager.__import_files(caption, last_folder, file_extensions, True)
    
    @staticmethod
    def __import_files(caption: str, last_folder: str, file_extensions: List[str], multiple_files: bool = False):
        imported_paths, file_extension = FileManager.get_file_paths(last_folder, file_extensions, caption, multiple_files)
        if not file_extension:
            return

        imported_data = list()
        if isinstance(imported_paths, list):
            for imported_path in imported_paths:
                imported_data.extend(FileManager.read_data_in_file(imported_path, use_first_sheet=False))

        else:
            imported_data.extend(FileManager.read_data_in_file(imported_paths, use_first_sheet=True))

        return imported_data

    @staticmethod
    def get_file_paths(file_extensions: List[str], caption: str = "Open file", multiple_files: bool = False, 
                       last_folder: str = None, open_file: bool = True):

        if last_folder is None:
            last_folder = os.path.expanduser("~")

        kwargs = dict()
        if platform.system() == "Linux":
            kwargs["options"] = QFileDialog.Option.DontUseNativeDialog

        str_extensions = "Files ("
        for extension in file_extensions:
            str_extensions += "*."
            str_extensions += extension
            str_extensions += " "
        
        str_extensions = str_extensions.strip()
        str_extensions += ")"

        imported_paths, file_extension = None, None

        if open_file:
            if multiple_files:
                imported_paths, file_extension = QFileDialog.getOpenFileNames(
                    None,
                    caption,
                    last_folder,
                    str_extensions,
                    **kwargs
                    )
        
            else:
                imported_paths, file_extension = QFileDialog.getOpenFileName(
                    None,
                    caption,
                    last_folder,
                    str_extensions,
                    **kwargs
                    )
        else:
            imported_paths, file_extensions = QFileDialog.getSaveFileName(
                None,
                caption,
                last_folder,
                str_extensions,
                **kwargs
            )
            
        return imported_paths, file_extension

    @staticmethod
    def read_data_in_file(file_path: str, use_first_sheet: bool = True) -> ImportedFile | list[ImportedFile]:
        import warnings

        output_data = list()

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore")
               
            sufix = Path(file_path).suffix
            filename = os.path.basename(file_path)

            if sufix in [".txt", ".dat", ".csv"]:
                try:
                    loaded_data = np.loadtxt(file_path, delimiter = ",")
                except:
                    loaded_data = FileManager.__load_text_file_data(file_path)

                loaded_data = FileManager.__remove_unnecesary_header_in_data(loaded_data)
                output_data.append(ImportedFile(loaded_data, filename, sufix, path=file_path))
                
            elif sufix in [".xls", ".xlsx"]:

                from polars import read_excel
                from openpyxl import load_workbook

                wb = load_workbook(file_path)
                
                for sheetname in wb.sheetnames:
                    max_cols = wb[sheetname].max_column 

                    for i in range(max_cols, 1, -1):
                        cols = list(range(i))

                        try:
                            sheet_data = read_excel(
                                                    file_path, 
                                                    sheet_name = sheetname,  
                                                    columns = cols,
                                                    engine = "openpyxl",
                                                    ).to_numpy()
                            break
                        except:
                            pass

                    sheet_data = FileManager.__remove_unnecesary_header_in_data(sheet_data)
                    output_data.append(ImportedFile(sheet_data, filename, sufix, sheetname, file_path))
                    if use_first_sheet:
                        break

            return output_data[0] if len(output_data) == 1 else output_data

    @staticmethod                      
    def __remove_unnecesary_header_in_data(data: np.ndarray) -> np.ndarray:
        filtered_data = [row for row in data if not isinstance(row[0], str)]
        return np.array(filtered_data, dtype=float)

    @staticmethod
    def __load_text_file_data(path: str):
        output_data = list()
        if isinstance(path, str):
            path = Path(path)

        with open(path, 'r') as file:
            for line in file.readlines():
                try:
                    modif_line = line.replace(",", " ").strip().split(" ")
                    line_values = [float(value) for value in modif_line if value != ""]
                except:
                    continue

                output_data.append(line_values)

        return np.array(output_data, dtype=float)

    @staticmethod
    def load_cfd_simulation_data_from_hdf_file(path: str | Path):
        simulation_data = dict()

        with h5py.File(path, 'r') as hf:

            # Read key data
            nodal_data = hf.get("nodal_data")
            variables = hf.get("variables")
            metadata = hf.get("metadata")

            # Save the nodal coordinates matrix
            simulation_data["nodal_area"] = np.array(nodal_data.get("nodal_area"), dtype=float)
            simulation_data["nodal_coordinates"] = np.array(nodal_data.get("coords"), dtype=float)

            # Save other attributes
            for attr_key in metadata.attrs:
                simulation_data[attr_key] = metadata.attrs[attr_key]

            # Calculate the start point from last revolution
            delta_theta = metadata.attrs["delta_theta"]
            steps_per_rev = int(360 / delta_theta)
            start = steps_per_rev + 1

            # filter the last revolution data for metadata
            for key, values in metadata.items():
                simulation_data[key] = values[-start:]

            # filter the last revolution data for variables
            for key, values in variables.items():
                simulation_data[key] = values[:, -start:]

        return simulation_data

    @staticmethod
    def load_spreadsheet_data_for_validation(path: str) -> dict:
        imported_results = dict()

        if not Path(path).exists():
            return imported_results
        
        from polars import read_excel
        from openpyxl import load_workbook

        wb = load_workbook(path)

        sheetnames = wb.sheetnames

        for sheetname in sheetnames:

            try:
                sheet_data = read_excel(
                                        path, 
                                        sheet_name = sheetname, 
                                        columns = [0, 1, 2]
                                        ).to_numpy()

            except:
                sheet_data = read_excel(
                                        path, 
                                        sheet_name = sheetname, 
                                        columns = [0, 1]
                                        ).to_numpy()
                
            filtered_data = [row_data for row_data in sheet_data if not isinstance(row_data[0], str)]
            sheet_data = np.array(filtered_data, dtype=float)

            imported_results[sheetname] = sheet_data

        return imported_results

    @staticmethod
    def export_text_data(export_path: str, exported_data: np.array, delimiter: str = ",", header: str = ""):
        export_path = Path(export_path)

        if not export_path.parent.exists():
            raise FileNotFoundError(f"The path {export_path.parent} does not exist")
        
        if export_path.suffix not in [".txt", ".dat", ".csv"]:
            raise ValueError(
            f"Invalid suffix {export_path.suffix}. Use .txt, .dat or .csv"
        )
        
        np.savetxt(str(export_path), exported_data, delimiter=delimiter, header=header)

    def export_spreadsheet_data(export_path: str, sheet_name: str, exported_data: PolarsDataFrame, row_indexes: bool = False):
        export_path = Path(export_path)

        if not export_path.parent.exists():
            raise FileNotFoundError(f"The path {export_path.parent} does not exist")
        
        if export_path.suffix not in [".xls", ".xlsx"]:
            raise ValueError(
            f"Invalid suffix {export_path.suffix}. Use .xls, or .xlsx"
        )

        from pandas import ExcelWriter

        with ExcelWriter(str(export_path)) as writer:
            exported_data.to_pandas().to_excel(writer, sheet_name=sheet_name, index=row_indexes)