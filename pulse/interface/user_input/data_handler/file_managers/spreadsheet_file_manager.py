from pulse.interface.user_input.data_handler.file_managers.io_handler import IOHandler
from pulse.interface.user_input.data_handler.imported_data.spreadsheet_data import SpreadsheetData

from polars import DataFrame as PolarsDataFrame
from pathlib import Path

import numpy as np


class SpreadsheetFileManager(IOHandler):

    def __init__(self):
        super().__init__()

    def read(self, file_path: str | Path) -> SpreadsheetData:
        file_path = Path(file_path)

        if file_path.suffix not in [".xls", ".xlsx"]:
            raise ValueError(
            f"Invalid suffix {file_path.suffix}. Use .xls, or .xlsx"
        )

        from polars import read_excel
        from openpyxl import load_workbook

        wb = load_workbook(file_path)

        imported_spreadsheet = SpreadsheetData(file_path.stem,
                                               file_path.suffix,
                                               str(file_path),
                                               wb.sheetnames
                                               )
        
        for sheetname in wb.sheetnames:
            max_cols = wb[sheetname].max_column 

            for i in range(max_cols, 1, -1):
                cols = list(range(i))

                try:
                    sheet_data = read_excel(
                                            str(file_path), 
                                            sheet_name = sheetname,  
                                            columns = cols,
                                            engine = "openpyxl",
                                            ).to_numpy()
                    break
                except:
                    pass

            sheet_data = self.__remove_unnecesary_header_in_data(sheet_data)
            imported_spreadsheet.data[sheetname] = sheet_data

        return imported_spreadsheet

    def save(self, file_path: str | Path, sheet_name: str, data: PolarsDataFrame, index_rows: bool = False):
        file_path = Path(file_path)

        if not file_path.parent.exists():
            raise FileNotFoundError(f"The path {file_path.parent} does not exist")
        
        if file_path.suffix not in [".xls", ".xlsx"]:
            raise ValueError(
            f"Invalid suffix {file_path.suffix}. Use .xls, or .xlsx"
        )

        from pandas import ExcelWriter

        with ExcelWriter(str(file_path)) as writer:
            data.to_pandas().to_excel(writer, sheet_name=sheet_name, index=index_rows)
    
    def __remove_unnecesary_header_in_data(self, data: np.ndarray) -> np.ndarray:
        filtered_data = [row for row in data if not isinstance(row[0], str)]
        return np.array(filtered_data, dtype=float)