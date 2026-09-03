from pathlib import Path

import numpy as np
from polars import DataFrame as PolarsDataFrame

from pulse.interface.user_input.data_handler.file_handlers.io_handler import IOHandler
from pulse.interface.user_input.data_handler.imported_data import (
    SpreadsheetData,
    SpreadsheetSheet,
)


class SpreadsheetFileHandler(IOHandler):
    EXTENSIONS = [".xls", ".xlsx"]

    @staticmethod
    def read(file_path: Path) -> SpreadsheetData:
        from openpyxl import load_workbook
        from polars import read_excel

        wb = load_workbook(file_path)

        imported_spreadsheet = SpreadsheetData(file_path)

        sheets = list()
        for sheetname in wb.sheetnames:
            max_cols = wb[sheetname].max_column

            for i in range(max_cols, 1, -1):
                cols = list(range(i))

                try:
                    sheet_data = read_excel(
                        str(file_path),
                        sheet_name=sheetname,
                        columns=cols,
                        engine="openpyxl",
                        has_header=False,
                        infer_schema_length=100,
                    ).to_numpy()
                    break
                except Exception:
                    pass

            sheet_data = SpreadsheetFileHandler._remove_unnecesary_header_in_data(sheet_data)
            sheets.append(SpreadsheetSheet(sheetname, sheet_data))

        imported_spreadsheet.sheets = sheets

        return imported_spreadsheet

    @staticmethod
    def save(file_path: str | Path, sheet_name: str, data: PolarsDataFrame, index_rows: bool = False, append=False):
        from pandas import ExcelWriter

        mode = "a" if append else "w"
        kwargs = {"if_sheet_exists": "replace"} if append else {}

        with ExcelWriter(str(file_path), mode=mode, **kwargs) as writer:
            data.to_pandas().to_excel(writer, sheet_name=sheet_name, index=index_rows)

    @staticmethod
    def _remove_unnecesary_header_in_data(data: np.ndarray) -> np.ndarray:
        filtered_data = [row for row in data if SpreadsheetFileHandler._is_valid_row(row)]
        return np.array(filtered_data, dtype=float)

    @staticmethod
    def _is_valid_row(row: np.ndarray) -> bool:
        try:
            float(row[0])
            return True
        except (ValueError, TypeError):
            return False