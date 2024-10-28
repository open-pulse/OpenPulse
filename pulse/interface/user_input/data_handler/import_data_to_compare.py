from PyQt5.QtWidgets import QDialog, QCheckBox, QFileDialog, QHBoxLayout, QLineEdit, QPushButton, QSpinBox, QTreeWidget, QTreeWidgetItem, QWidget
from PyQt5.QtGui import QCloseEvent
from PyQt5.QtCore import Qt
from PyQt5 import uic

from pulse import app, UI_DIR
from pulse.interface.user_input.project.print_message import PrintMessageInput

import os
import numpy as np
from pathlib import Path

window_title_1 = "Error"
window_title_2 = "Warning"

class ImportDataToCompare(QDialog):
    def __init__(self, plotter, *args, **kwargs):
        super().__init__(*args, **kwargs)

        ui_path = UI_DIR / "data_handler/import_data_to_compare.ui"
        uic.loadUi(ui_path, self)
        
        self.plotter = plotter

        self.main_window = app().main_window

        self._config_window()
        self._initialize()
        self._define_qt_variables()
        self._create_connections()
        self._config_widgets()
        
    def _config_window(self):    
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)
        self.setWindowIcon(app().main_window.pulse_icon)
        self.setWindowTitle("OpenPulse")

    def _initialize(self):

        self.keep_window_open = True
        self.imported_data = None

        self.imported_results = dict()
        self.ids_to_checkBox = dict()
        self.checkButtons_state = dict()

        self.colors = [ [0,0,0],
                        [1,0,0],
                        [1,0,1],
                        [0,1,1],
                        [0.75,0.75,0.75],
                        [0.5, 0.5, 0.5],
                        [0.25, 0.25, 0.25] ]

    def _define_qt_variables(self):

        # CheckBox
        self.checkBox_skiprows: QCheckBox

        # LineEdit
        self.lineEdit_import_results_path: QLineEdit
        self.lineEdit_import_results_path.setDisabled(True)

        # PushButton
        self.pushButton_add_imported_data_to_plot: QPushButton
        self.pushButton_cancel: QPushButton
        self.pushButton_reset_imported_data: QPushButton
        self.pushButton_search_file_to_import: QPushButton

        # SpinBox
        self.spinBox_skiprows: QSpinBox

        # TreeWidget
        self.treeWidget_import_text_files: QTreeWidget
        self.treeWidget_import_sheet_files: QTreeWidget

    def _create_connections(self):
        #
        self.checkBox_skiprows.clicked.connect(self.update_skiprows_visibility)
        #
        self.pushButton_search_file_to_import.clicked.connect(self.choose_path_to_import_results)
        self.pushButton_cancel.clicked.connect(self.close)
        self.pushButton_reset_imported_data.clicked.connect(self.reset_imported_data)
        self.pushButton_add_imported_data_to_plot.clicked.connect(self.add_imported_data_to_plot)
        #
        self.update_skiprows_visibility()
        
    def _config_widgets(self):

        widths_1 = [320, 60]
        for i, width in enumerate(widths_1):
            self.treeWidget_import_text_files.setColumnWidth(i, width)

        widths_2 = [180, 180, 60]
        for i, width in enumerate(widths_2):
            self.treeWidget_import_sheet_files.setColumnWidth(i, width)

    def update_skiprows_visibility(self):
        self.spinBox_skiprows.setDisabled(not self.checkBox_skiprows.isChecked())

    def choose_path_to_import_results(self):

        path = app().config.get_last_folder_for("imported data folder")
        if path is None:
            folder_path = os.path.expanduser("~")
        else:
            folder_path = path

        imported_path, check = QFileDialog.getOpenFileName( None, 
                                                            'Open file', 
                                                            folder_path, 
                                                            'Files (*.csv *.dat *.txt *.xlsx *.xls)' )

        if not check:
            return

        app().config.write_last_folder_path_in_file("imported data folder", imported_path)

        self.import_name = os.path.basename(imported_path)
        self.lineEdit_import_results_path.setText(imported_path)

        self.import_results(imported_path)
        self.update_treeWidget_info()

    def import_results(self, imported_path: str):

        from pandas import read_excel
        from openpyxl import load_workbook

        try:

            message = ""

            run = True
            if self.checkBox_skiprows.isChecked():
                skiprows = self.spinBox_skiprows.value()
            else:
                skiprows = 0
            maximum_lines_to_skip = 100
            
            while run:
                try:
                    sufix = Path(imported_path).suffix
                    filename = os.path.basename(imported_path)
                    if sufix in [".txt", ".dat", ".csv"]:
                        loaded_data = np.loadtxt(imported_path, 
                                                 delimiter = ",", 
                                                 skiprows = skiprows)
                        key = self.get_data_index()
                        self.imported_results[key] = {  "data" : loaded_data,
                                                        "filename" : filename,
                                                        "extension" : sufix  }

                    elif sufix in [".xls", ".xlsx"]:
                        wb = load_workbook(imported_path)
                        sheetnames = wb.sheetnames
                        for sheetname in sheetnames:

                            try:
                                sheet_data = read_excel(
                                                        imported_path, 
                                                        sheet_name = sheetname, 
                                                        header = skiprows, 
                                                        usecols = [0,1,2]
                                                        ).to_numpy()
                            except:
                                sheet_data = read_excel(
                                                        imported_path, 
                                                        sheet_name = sheetname, 
                                                        header = skiprows, 
                                                        usecols = [0,1]
                                                        ).to_numpy()

                            key = self.get_data_index()
                            self.imported_results[key] = {  "data" : sheet_data,
                                                            "filename" : filename,
                                                            "sheetname" : sheetname,
                                                            "extension" : sufix  }

                    self.spinBox_skiprows.setValue(int(skiprows))
                    run = False

                except:
                    skiprows += 1
                    if skiprows >= maximum_lines_to_skip:
                        run = False
                        title = "Error while loading data from file"
                        message = "The maximum number of rows to skip has been reached and no valid data has "
                        message += "been found. Please, verify the data in the imported file to proceed."
                        message += "Maximum number of header rows: 100"

        except Exception as log_error:
            title = "Error while loading data from file"
            message = str(log_error)
            return
        
        if message != "":
            PrintMessageInput([window_title_1, title, message])

    def update_treeWidget_info(self):
        self.cache_checkButtons_state()
        self.treeWidget_import_text_files.clear()
        self.treeWidget_import_sheet_files.clear()
        #
        if len(self.imported_results) > 0:
            for i, (id, data) in enumerate(self.imported_results.items()):
                # Creates the QCheckButtons to control data to be plotted
                self.ids_to_checkBox[id] = QCheckBox()
                self.ids_to_checkBox[id].setStyleSheet("margin-left:40%; margin-right:50%;")

                if id in self.checkButtons_state.keys():
                    self.ids_to_checkBox[id].setChecked(self.checkButtons_state[id])

                if "sheetname" in data.keys():
                    _item = QTreeWidgetItem([str(data["filename"]), str(data["sheetname"])])
                    self.treeWidget_import_sheet_files.addTopLevelItem(_item)
                    self.treeWidget_import_sheet_files.setItemWidget(_item, 2, self.ids_to_checkBox[id])
                else:
                    _item = QTreeWidgetItem([str(data["filename"])])
                    self.treeWidget_import_text_files.addTopLevelItem(_item)
                    self.treeWidget_import_text_files.setItemWidget(_item, 1, self.ids_to_checkBox[id])                  

                for i in range(5):
                    _item.setTextAlignment(i, Qt.AlignCenter)

    def get_data_index(self):
        index = 1
        run = True
        while run:
            # if index in self.plotter.model_results_data.keys() or index in self.imported_results.keys():
            if index in self.imported_results.keys():
                index += 1
            else:
                key = index
                run = False
                
        return key
    
    def join_imported_data(self):
        j = 0
        imported_results_data = dict()
        for id, checkBox in self.ids_to_checkBox.items():
            temp_dict = dict()
            if checkBox.isChecked():

                if id < len(self.colors):
                    color = self.colors[j]
                    j += 1
                else:
                    color = np.random.randint(0,255,3)/255

                data = self.imported_results[id]["data"]
                cols = data.shape[1]
                x_values = data[:, 0]
                if cols == 2:
                    y_values = data[:, 1]
                else:
                    y_values = data[:, 1] + 1j*data[:, 2]

                if "sheetname" in self.imported_results[id].keys():
                    sheetname = self.imported_results[id]["sheetname"]
                    legend_label = f"{sheetname}"
                else:
                    legend_label = self.imported_results[id]["filename"]

                temp_dict = {   "type" : "imported_data",
                                "x_data" : x_values,
                                "y_data" : y_values,
                                "x_label" : "Frequency [Hz]",
                                "y_label" : "Nodal response",
                                "legend" : legend_label,
                                "unit" : "",
                                "title" : "",
                                "color" : color,
                                "linestyle" : "--"   }

                key = (id)
                imported_results_data[key] = temp_dict

        self.plotter._set_imported_results_data_to_plot(imported_results_data)

    def cache_checkButtons_state(self):
        self.checkButtons_state = dict()
        for key, check in self.ids_to_checkBox.items():
            self.checkButtons_state[key] = check.isChecked()

    def reset_imported_data(self):
        self.lineEdit_import_results_path.setText("")
        self.treeWidget_import_sheet_files.clear()
        self.treeWidget_import_text_files.clear()
        self._initialize()

    def add_imported_data_to_plot(self):
        self.join_imported_data()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.add_imported_data_to_plot()
        elif event.key() == Qt.Key_Escape:
            self.close()

    def closeEvent(self, a0: QCloseEvent | None) -> None:

        # if self.exporter is not None:
        #     self.exporter.close()

        # if self.importer is not None:
        #     self.importer.close()

        self.keep_window_open = False
        return super().closeEvent(a0)