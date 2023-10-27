from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from PyQt5 import uic
from pathlib import Path

import os
import numpy as np
import pandas as pd
import openpyxl

from data.user_input.project.printMessageInput import PrintMessageInput

def get_icons_path(filename):
    path = f"data/icons/{filename}"
    if os.path.exists(path):
        return str(Path(path))

class ImportDataToCompare(QDialog):
    def __init__(self, plotter, *args, **kwargs):
        super().__init__(*args, **kwargs)

        uic.loadUi(Path('data/user_input/ui_files/data_handler/import_data_to_compare.ui'), self)

        self.plotter = plotter

        self._config_window()
        self._load_icons()
        self._reset_variables()
        self._define_and_configure_Qt_variables()
        self._create_connections()
        self.exec()

    def _config_window(self):
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)
        self.setWindowTitle("Import data to compare")

    def _load_icons(self):
        self.icon = QIcon(get_icons_path('pulse.png'))
        self.search_icon = QIcon(get_icons_path('searchFile.png'))
        self.setWindowIcon(self.icon)

    def _reset_variables(self):
        self.userPath = os.path.expanduser('~')
        self.imported_path = ""
        self.imported_data = None
        self.imported_results = dict()
        self.ids_to_checkBox = dict()
        self.checkButtons_state = dict()
        self.colors = [ [0,0,0],
                        [0,1,0],
                        [1,1,0],
                        [0,1,1],
                        [1,0,1],
                        [0.75,0.75,0.75],
                        [0.5, 0.5, 0.5],
                        [0.25, 0.25, 0.25],
                        [0,0,1] ]

    def _define_and_configure_Qt_variables(self):
        # CheckBox
        self.checkBox_skiprows = self.findChild(QCheckBox, "checkBox_skiprows")
        # LineEdit
        self.lineEdit_import_results_path = self.findChild(QLineEdit, 'lineEdit_import_results_path')
        self.lineEdit_import_results_path.setDisabled(True)
        # PushButton
        self.pushButton_add_imported_data_to_plot = self.findChild(QPushButton, 'pushButton_add_imported_data_to_plot')
        self.pushButton_reset_imported_data = self.findChild(QPushButton, 'pushButton_reset_imported_data')
        self.pushButton_search_file_to_import = self.findChild(QPushButton, 'pushButton_search_file_to_import')
        self.pushButton_search_file_to_import.setIcon(self.search_icon)
        # SpinBox
        self.spinBox_skiprows = self.findChild(QSpinBox, 'spinBox_skiprows')
        # TreeWidget
        self.treeWidget_import_text_files = self.findChild(QTreeWidget, "treeWidget_import_text_files")
        self.treeWidget_import_sheet_files = self.findChild(QTreeWidget, "treeWidget_import_sheet_files")

        widths_1 = [320, 60]
        for i, width in enumerate(widths_1):
            self.treeWidget_import_text_files.setColumnWidth(i, width)

        widths_2 = [180, 150, 60]
        for i, width in enumerate(widths_2):
            self.treeWidget_import_sheet_files.setColumnWidth(i, width)

    def _create_connections(self):
        self.checkBox_skiprows.clicked.connect(self.update_skiprows_visibility)
        self.pushButton_search_file_to_import.clicked.connect(self.choose_path_import_results)
        self.pushButton_reset_imported_data.clicked.connect(self.reset_imported_data)
        self.pushButton_add_imported_data_to_plot.clicked.connect(self.add_imported_data_to_plot)
        self.update_skiprows_visibility()
        
    def update_skiprows_visibility(self):
        self.spinBox_skiprows.setDisabled(not self.checkBox_skiprows.isChecked())

    def choose_path_import_results(self):
        if self.imported_path == "":
            _path = self.userPath
        else:
            _path = os.path.dirname(self.imported_path)

        self.imported_path, _ = QFileDialog.getOpenFileName(None, 'Open file', _path, 'Files (*.csv *.dat *.txt *.xlsx *.xls)')
        self.import_name = os.path.basename(self.imported_path)
        self.lineEdit_import_results_path.setText(self.imported_path)
        
        if self.imported_path != "":
            if os.path.exists(self.imported_path):
                self.import_results()
                self.update_treeWidget_info()

    def import_results(self):
        
        try:

            window_title = "ERROR MESSAGE"
            message = ""

            run = True
            if self.checkBox_skiprows.isChecked():
                skiprows = self.spinBox_skiprows.value()
            else:
                skiprows = 0
            maximum_lines_to_skip = 100
            
            while run:
                try:
                    sufix = Path(self.imported_path).suffix
                    filename = os.path.basename(self.imported_path)
                    if sufix in [".txt", ".dat", ".csv"]:
                        loaded_data = np.loadtxt(self.imported_path, 
                                                 delimiter = ",", 
                                                 skiprows = skiprows)
                        key = self.get_data_index()
                        self.imported_results[key] = {  "data" : loaded_data,
                                                        "filename" : filename,
                                                        "extension" : sufix  }

                    elif sufix in [".xls", ".xlsx"]:
                        wb = openpyxl.load_workbook(self.imported_path)
                        sheetnames = wb.sheetnames
                        for sheetname in sheetnames:

                            try:
                                sheet_data = pd.read_excel(self.imported_path, 
                                                        sheet_name = sheetname, 
                                                        header = skiprows, 
                                                        usecols = [0,1,2]).to_numpy()
                            except:
                                sheet_data = pd.read_excel(self.imported_path, 
                                                        sheet_name = sheetname, 
                                                        header = skiprows, 
                                                        usecols = [0,1]).to_numpy()

                            key = self.get_data_index()
                            self.imported_results[key] = {  "data" : sheet_data,
                                                            "filename" : filename,
                                                            "sheetname" : sheetname,
                                                            "extension" : sufix  }

                    self.spinBox_skiprows.setValue(int(skiprows))
                    run = False

                except:
                    skiprows += 1
                    if skiprows >= maximum_lines_to_skip:
                        run = False
                        title = "Error while loading data from file"
                        message = "The maximum number of rows to skip has been reached and no valid data has "
                        message += "been found. Please, verify the data in the imported file to proceed."
                        message += "Maximum number of header rows: 100"

        except Exception as log_error:
            title = "Error while loading data from file"
            message = str(log_error)
            return
        
        if message != "":
            PrintMessageInput([title, message, window_title])

    def update_treeWidget_info(self):
        self.cache_checkButtons_state()
        self.treeWidget_import_text_files.clear()
        self.treeWidget_import_sheet_files.clear()
        #
        if len(self.imported_results) > 0:
            for i, (id, data) in enumerate(self.imported_results.items()):
                # Creates the QCheckButtons to control data to be plotted
                self.ids_to_checkBox[id] = QCheckBox()
                self.ids_to_checkBox[id].setStyleSheet("margin-left:40%; margin-right:50%;")

                if id in self.checkButtons_state.keys():
                    self.ids_to_checkBox[id].setChecked(self.checkButtons_state[id])

                if "sheetname" in data.keys():
                    _item = QTreeWidgetItem([str(data["filename"]), str(data["sheetname"])])
                    self.treeWidget_import_sheet_files.addTopLevelItem(_item)
                    self.treeWidget_import_sheet_files.setItemWidget(_item, 2, self.ids_to_checkBox[id])
                else:
                    _item = QTreeWidgetItem([str(data["filename"])])
                    self.treeWidget_import_text_files.addTopLevelItem(_item)
                    self.treeWidget_import_text_files.setItemWidget(_item, 1, self.ids_to_checkBox[id])                  

                for i in range(5):
                    _item.setTextAlignment(i, Qt.AlignCenter)

    def get_data_index(self):
        index = 1
        run = True
        while run:
            if index in self.plotter.data_to_plot.keys() or index in self.imported_results.keys():
                index += 1
            else:
                key = index
                run = False
                
        return key
    
    def join_imported_data(self):
        j = 0
        for id, checkBox in self.ids_to_checkBox.items():
            temp_dict = dict()
            if checkBox.isChecked():

                if id < len(self.colors):
                    color = self.colors[j]
                    j += 1
                else:
                    color = np.random.randint(0,255,3)/255

                data = self.imported_results[id]["data"]
                cols = data.shape[1]
                x_values = data[:, 0]
                if cols == 2:
                    y_values = data[:, 1]
                else:
                    y_values = data[:, 1] + 1j*data[:, 2]

                if "sheetname" in self.imported_results[id].keys():
                    sheetname = self.imported_results[id]["sheetname"]
                    legend_label = f"{sheetname}"
                else:
                    legend_label = self.imported_results[id]["filename"]

                temp_dict = {   "type" : "imported_data",
                                "x_data" : x_values,
                                "y_data" : y_values,
                                "x_label" : "Frequency [Hz]",
                                "y_label" : "Nodal response",
                                "legend" : legend_label,
                                "unit" : "",
                                "title" : "",
                                "color" : color,
                                "linestyle" : "--"   }

                self.plotter.data_to_plot[id] = temp_dict

    def cache_checkButtons_state(self):
        self.checkButtons_state = dict()
        for key, check in self.ids_to_checkBox.items():
            self.checkButtons_state[key] = check.isChecked()

    def reset_imported_data(self):
        self.lineEdit_import_results_path.setText("")
        self.treeWidget_import_sheet_files.clear()
        self.treeWidget_import_text_files.clear()
        self._reset_variables()

    def add_imported_data_to_plot(self):
        self.join_imported_data()
        self.plotter.plot_data_in_freq_domain()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.add_imported_data_to_plot()
        elif event.key() == Qt.Key_Escape:
            self.close()