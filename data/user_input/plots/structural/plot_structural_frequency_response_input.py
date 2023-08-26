from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from PyQt5 import uic
from pathlib import Path

import os
import numpy as np
import pandas as pd
import openpyxl

from pulse.postprocessing.plot_structural_data import get_structural_frf
from data.user_input.project.printMessageInput import PrintMessageInput
from data.user_input.plots.general.frequency_response_plotter import FrequencyResponsePlotter
from data.user_input.data_handler.import_data_to_compare import ImportDataToCompare

def get_icons_path(filename):
    path = f"data/icons/{filename}"
    if os.path.exists(path):
        return str(Path(path))

window_title1 = "ERROR MESSAGE"
window_title2 = "WARNING MESSAGE"

class PlotStructuralFrequencyResponseInput(QDialog):
    def __init__(self, project, opv, analysisMethod, solution, *args, **kwargs):
        super().__init__(*args, **kwargs)

        uic.loadUi(Path('data/user_input/ui_files/plots_/results_/structural_/plot_structural_frequency_response.ui'), self)

        self.icon = QIcon(get_icons_path('pulse.png'))
        self.search_icon = QIcon(get_icons_path('searchFile.png'))
        self.setWindowIcon(self.icon)

        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)

        self.opv = opv
        self.opv.setInputObject(self)
        self.list_node_IDs = self.opv.getListPickedPoints()

        self.projec = project
        self.preprocessor = project.preprocessor
        self.before_run = project.get_pre_solution_model_checks()
        self.nodes = self.preprocessor.nodes
        
        self.analysisMethod = analysisMethod
        self.frequencies = project.frequencies
        self.solution = solution

        self._reset_variables()
        self._define_and_configure_Qt_variables()
        self._create_connections()
        self.writeNodes(self.list_node_IDs)
        self.exec()

    def _reset_variables(self):
        self.userPath = os.path.expanduser('~')
        self.imported_path = ""
        self.save_path = ""
        self.node_ID = 0
        self.imported_data = None
        self.localDof = None
        self.imported_results = dict()
        self.ids_to_checkBox = dict()
        self.checkButtons_state = dict()
        self.colors = [ [0,0,0],
                        [0,1,0],
                        [1,1,0],
                        [0,1,1],
                        [1,0,1],
                        [0.75,0.75,0.75],
                        [0.5, 0.5, 0.5],
                        [0.25, 0.25, 0.25],
                        [0,0,1] ]

    def _define_and_configure_Qt_variables(self):
        # CheckBox
        self.checkBox_skiprows = self.findChild(QCheckBox, "checkBox_skiprows")
        # LineEdit
        self.lineEdit_node_id = self.findChild(QLineEdit, 'lineEdit_node_id')
        self.lineEdit_file_name = self.findChild(QLineEdit, 'lineEdit_file_name')
        self.lineEdit_import_results_path = self.findChild(QLineEdit, 'lineEdit_import_results_path')
        self.lineEdit_save_results_path = self.findChild(QLineEdit, 'lineEdit_save_results_path')
        self.lineEdit_import_results_path.setDisabled(True)
        # PushButton
        self.pushButton_plot_frequency_response = self.findChild(QPushButton, 'pushButton_plot_frequency_response')
        self.pushButton_search_file_to_import = self.findChild(QPushButton, 'pushButton_search_file_to_import')
        self.pushButton_choose_folder_export = self.findChild(QPushButton, 'pushButton_choose_folder_export')
        self.pushButton_export_results = self.findChild(QPushButton, 'pushButton_export_results')
        self.pushButton_reset_filename = self.findChild(QPushButton, 'pushButton_reset_filename')
        self.pushButton_search_file_to_import.setIcon(self.search_icon)
        self.pushButton_choose_folder_export.setIcon(self.search_icon)
        # RadioButton
        self.radioButton_ux = self.findChild(QRadioButton, 'radioButton_ux')
        self.radioButton_uy = self.findChild(QRadioButton, 'radioButton_uy')
        self.radioButton_uz = self.findChild(QRadioButton, 'radioButton_uz')
        self.radioButton_rx = self.findChild(QRadioButton, 'radioButton_rx')
        self.radioButton_ry = self.findChild(QRadioButton, 'radioButton_ry')
        self.radioButton_rz = self.findChild(QRadioButton, 'radioButton_rz')
        self.radioButton_none_diff = self.findChild(QRadioButton, 'radioButton_none_diff')
        self.radioButton_single_diff = self.findChild(QRadioButton, 'radioButton_single_diff')
        self.radioButton_double_diff = self.findChild(QRadioButton, 'radioButton_double_diff')
        # SpinBox
        self.spinBox_skiprows = self.findChild(QSpinBox, 'spinBox_skiprows')
        # TabWidget
        self.tabWidget_plot_results = self.findChild(QTabWidget, "tabWidget_plot_results")
        self.tab_plot = self.tabWidget_plot_results.findChild(QWidget, "tab_plot")
        # TreeWidget
        self.treeWidget_import_text_files = self.findChild(QTreeWidget, "treeWidget_import_text_files")
        self.treeWidget_import_sheet_files = self.findChild(QTreeWidget, "treeWidget_import_sheet_files")

        widths_1 = [320, 60]
        for i, width in enumerate(widths_1):
            self.treeWidget_import_text_files.setColumnWidth(i, width)

        widths_2 = [180, 150, 60]
        for i, width in enumerate(widths_2):
            self.treeWidget_import_sheet_files.setColumnWidth(i, width)

    def _create_connections(self):
        self.checkBox_skiprows.clicked.connect(self.update_skiprows_visibility)
        self.pushButton_search_file_to_import.clicked.connect(self.choose_path_import_results)
        self.pushButton_choose_folder_export.clicked.connect(self.choose_path_export_results)
        self.pushButton_export_results.clicked.connect(self.export_results)
        self.pushButton_plot_frequency_response.clicked.connect(self.check_inputs_and_plot)
        self.pushButton_reset_filename.clicked.connect(self.reset_filename)
        self.update_skiprows_visibility()

    def reset_filename(self):
        self.lineEdit_file_name.setText("")
        self.lineEdit_file_name.setFocus()

    def update_skiprows_visibility(self):
        self.spinBox_skiprows.setDisabled(not self.checkBox_skiprows.isChecked())

    def reset_imported_data(self):
        self._reset_variables()
        self.update_treeWidget_info()
        self.lineEdit_import_results_path.setText("")
        title = "Information"
        message = "The plot data has been reseted."
        PrintMessageInput([title, message, window_title2])
    
    def writeNodes(self, list_node_ids):
        text = ""
        for node in list_node_ids:
            text += "{}, ".format(node)
        self.lineEdit_node_id.setText(text)

    def update(self):
        self.list_node_IDs = self.opv.getListPickedPoints()
        if self.list_node_IDs != []:
            self.writeNodes(self.list_node_IDs)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.check_inputs_and_plot()
        elif event.key() == Qt.Key_Escape:
            self.close()

    def choose_path_import_results(self):
        if self.imported_path == "":
            _path = self.userPath
        else:
            _path = os.path.dirname(self.imported_path)

        self.imported_path, _ = QFileDialog.getOpenFileName(None, 'Open file', _path, 'Files (*.csv *.dat *.txt *.xlsx *.xls)')
        self.import_name = os.path.basename(self.imported_path)
        self.lineEdit_import_results_path.setText(self.imported_path)
        
        if self.imported_path != "":
            if os.path.exists(self.imported_path):
                self.import_results()
                self.update_treeWidget_info()

    def get_data_index(self):
        index = 1
        run = True
        while run:
            if index not in self.imported_results.keys():
                key = index
                run = False
            else:
                index += 1
        return key

    def import_results(self):
        try:
            message = ""
            run = True
            if self.checkBox_skiprows.isChecked():
                skiprows = self.spinBox_skiprows.value()
            else:
                skiprows = 0
            maximum_lines_to_skip = 100
            
            while run:
                try:
                    sufix = Path(self.imported_path).suffix
                    filename = os.path.basename(self.imported_path)
                    if sufix in [".txt", ".dat", ".csv"]:
                        loaded_data = np.loadtxt(self.imported_path, 
                                                 delimiter = ",", 
                                                 skiprows = skiprows)
                        key = self.get_data_index()
                        self.imported_results[key] = {  "data" : loaded_data,
                                                        "filename" : filename,
                                                        "extension" : sufix  }

                    elif sufix in [".xls", ".xlsx"]:
                        wb = openpyxl.load_workbook(self.imported_path)
                        sheetnames = wb.sheetnames
                        for sheetname in sheetnames:
                            sheet_data = pd.read_excel(self.imported_path, 
                                                       sheet_name = sheetname, 
                                                       header = skiprows, 
                                                       usecols = [0,1,2]).to_numpy()
                            key = self.get_data_index()
                            self.imported_results[key] = {  "data" : sheet_data,
                                                            "filename" : filename,
                                                            "sheetname" : sheetname,
                                                            "extension" : sufix  }

                    self.spinBox_skiprows.setValue(int(skiprows))
                    run = False

                except:
                    skiprows += 1
                    if skiprows >= maximum_lines_to_skip:
                        run = False
                        title = "Error while loading data from file"
                        message = "The maximum number of rows to skip has been reached and no valid data has "
                        message += "been found. Please, verify the data in the imported file to proceed."
                        message += "Maximum number of header rows: 100"

        except Exception as log_error:
            title = "Error while loading data from file"
            message = str(log_error)
            return
        
        if message != "":
            PrintMessageInput([title, message, window_title1])

    def update_treeWidget_info(self):
        self.cache_checkButtons_state()
        self.treeWidget_import_text_files.clear()
        self.treeWidget_import_sheet_files.clear()
        #
        if len(self.imported_results) > 0:
            for i, (id, data) in enumerate(self.imported_results.items()):
                # Creates the QCheckButtons to control data to be plotted
                self.ids_to_checkBox[id] = QCheckBox()
                self.ids_to_checkBox[id].setStyleSheet("margin-left:40%; margin-right:50%;")

                if id in self.checkButtons_state.keys():
                    self.ids_to_checkBox[id].setChecked(self.checkButtons_state[id])

                if "sheetname" in data.keys():
                    _item = QTreeWidgetItem([str(data["filename"]), str(data["sheetname"])])
                    self.treeWidget_import_sheet_files.addTopLevelItem(_item)
                    self.treeWidget_import_sheet_files.setItemWidget(_item, 2, self.ids_to_checkBox[id])
                else:
                    _item = QTreeWidgetItem([str(data["filename"])])
                    self.treeWidget_import_text_files.addTopLevelItem(_item)
                    self.treeWidget_import_text_files.setItemWidget(_item, 1, self.ids_to_checkBox[id])                  

                for i in range(5):
                    _item.setTextAlignment(i, Qt.AlignCenter)

    def cache_checkButtons_state(self):
        self.checkButtons_state = {}
        for key, check in self.ids_to_checkBox.items():
            self.checkButtons_state[key] = check.isChecked()

    def choose_path_export_results(self):
        self.save_path = QFileDialog.getExistingDirectory(None, 'Choose a folder to export the results', self.userPath)
        self.save_name = os.path.basename(self.save_path)
        self.lineEdit_save_results_path.setText(str(self.save_path))

    def check_inputs_and_plot(self, export=False):
        lineEdit_node_id = self.lineEdit_node_id.text()
        stop, self.node_ID = self.before_run.check_input_NodeID(lineEdit_node_id, single_ID=True)
        if stop:
            return True

        self.localDof = None
        if self.radioButton_single_diff.isChecked():
            _unit_label = "m/s"
        elif self.radioButton_double_diff.isChecked():
            _unit_label = "m/s²"
        else:
            _unit_label = "m"    

        if self.radioButton_ux.isChecked():
            self.localDof = 0
            self.localdof_label = "Ux"
            self.unit_label = _unit_label

        if self.radioButton_uy.isChecked():
            self.localDof = 1
            self.localdof_label = "Uy"
            self.unit_label = _unit_label

        if self.radioButton_uz.isChecked():
            self.localDof = 2
            self.localdof_label = "Uz"
            self.unit_label = _unit_label
 
        if self.radioButton_rx.isChecked():
            self.localDof = 3
            self.localdof_label = "Rx"
            self.unit_label = _unit_label

        if self.radioButton_ry.isChecked():
            self.localDof = 4
            self.localdof_label = "Ry"
            self.unit_label = _unit_label

        if self.radioButton_rz.isChecked():
            self.localDof = 5
            self.localdof_label = "Rz"
            self.unit_label = _unit_label

        if self.radioButton_single_diff.isChecked():
            _unit_label = "rad/s"
        elif self.radioButton_double_diff.isChecked():
            _unit_label = "rad/s²"
        else:
            _unit_label = "rad"

        if not export:
            self.call_plotter()

        return False

    def export_results(self):
        
        if self.lineEdit_file_name.text() != "":
            if self.save_path != "":
                self.export_path_folder = self.save_path + "/"
            else:
                title = "None folder selected"
                message = "Plese, choose a folder before trying export the results."
                PrintMessageInput([title, message, window_title1])
                return
        else:
            title = "Empty file name"
            message = "Inform a file name before trying export the results."
            PrintMessageInput([title, message, window_title1])
            return
        
        if self.check_inputs_and_plot(export=True):
            return

        freq = self.frequencies
        self.export_path = self.export_path_folder + self.lineEdit_file_name.text() + ".dat"
        response = self.get_response()
        
        header = ("Frequency[Hz], Real part [{}], Imaginary part [{}], Absolute [{}]").format(self.unit_label, self.unit_label, self.unit_label)
        data_to_export = np.array([freq, np.real(response), np.imag(response), np.abs(response)]).T      
            
        np.savetxt(self.export_path, data_to_export, delimiter=",", header=header)
        title = "Information"
        message = "The results have been exported."
        PrintMessageInput([title, message, window_title2])

    def get_response(self):
        response = get_structural_frf(self.preprocessor, self.solution, self.node_ID, self.localDof)
        if self.radioButton_single_diff.isChecked():
            output_data = response*(1j*2*np.pi)*self.frequencies
        elif self.radioButton_double_diff.isChecked():
            output_data = response*((1j*2*np.pi*self.frequencies)**2)
        else:
            output_data = response
        return output_data

    def call_plotter(self):
        self.join_model_data()
        self.join_imported_data()
        self.plotter = FrequencyResponsePlotter()
        self.plotter._set_data_to_plot(self.data_to_plot)

    def join_model_data(self):
        self.data_to_plot = dict()
        self.title = "Structural frequency response - {}".format(self.analysisMethod)
        legend_label = "Response {} at node {}".format(self.localdof_label, self.node_ID)
        self.data_to_plot[0] = {   "type" : "model results",
                                    "x_data" : self.frequencies,
                                    "y_data" : self.get_response(),
                                    "x_label" : "Frequency [Hz]",
                                    "y_label" : "Nodal response",
                                    "legend" : legend_label,
                                    "unit" : self.unit_label,
                                    "title" : self.title,
                                    "color" : [0,0,1],
                                    "linestyle" : "-"   }

    def join_imported_data(self):
        j = 0
        for id, checkBox in self.ids_to_checkBox.items():
            temp_dict = dict()
            if checkBox.isChecked():

                if id < len(self.colors):
                    color = self.colors[j]
                    j += 1
                else:
                    color = np.random.randint(0,255,3)/255

                data = self.imported_results[id]["data"]
                x_values = data[:, 0]
                y_values = data[:, 1] + 1j*data[:, 2]

                if "sheetname" in self.imported_results[id].keys():
                    sheetname = self.imported_results[id]["sheetname"]
                    legend_label = f"{sheetname}"
                else:
                    legend_label = self.imported_results[id]["filename"]

                temp_dict = {   "type" : "imported_data",
                                "x_data" : x_values,
                                "y_data" : y_values,
                                "x_label" : "Frequency [Hz]",
                                "y_label" : "Nodal response",
                                "legend" : legend_label,
                                "unit" : self.unit_label,
                                "title" : self.title,
                                "color" : color,
                                "linestyle" : "--"   }

                self.data_to_plot[id] = temp_dict