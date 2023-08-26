from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
from PyQt5 import uic
from pathlib import Path

import os
import numpy as np
import pandas as pd
import openpyxl

from pulse.postprocessing.plot_acoustic_data import get_acoustic_frf
from data.user_input.project.printMessageInput import PrintMessageInput
from data.user_input.plots.general.frequency_response_plotter import FrequencyResponsePlotter


window_title1 = "ERROR MESSAGE"
window_title2 = "WARNING MESSAGE"


class PlotAcousticFrequencyResponseInput(QDialog):
    def __init__(self, project, opv, analysisMethod, solution, *args, **kwargs):
        super().__init__(*args, **kwargs)

        uic.loadUi(Path('data/user_input/ui_files/plots_/results_/acoustic_/plot_acoustic_frequency_response_input.ui'), self)

        icons_path = str(Path('data/icons/pulse.png'))
        self.icon = QIcon(icons_path)
        self.setWindowIcon(self.icon) 
        
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.setWindowModality(Qt.WindowModal)

        self.opv = opv
        self.opv.setInputObject(self)
        self.list_node_IDs = self.opv.getListPickedPoints()

        self.projec = project
        self.preprocessor = project.preprocessor
        self.before_run = project.get_pre_solution_model_checks()

        self.nodes = self.preprocessor.nodes
        self.analysisMethod = analysisMethod
        self.frequencies = project.frequencies
        self.solution = solution

        self._reset_variables()
        self._define_qt_variables()
        self._create_connections()
        self.writeNodes(self.list_node_IDs)
        self.exec()


    def _reset_variables(self):
        self.userPath = os.path.expanduser('~')
        self.save_path = ""
        self.node_ID = 0
        self.imported_data = None


    def _define_qt_variables(self):
        # QCheckBox
        self.checkBox_dB = self.findChild(QCheckBox, 'checkBox_dB')
        # QLineEdit
        self.lineEdit_FileName = self.findChild(QLineEdit, 'lineEdit_FileName')
        self.lineEdit_ImportResultsPath = self.findChild(QLineEdit, 'lineEdit_ImportResultsPath')
        self.lineEdit_SaveResultsPath = self.findChild(QLineEdit, 'lineEdit_SaveResultsPath')
        self.lineEdit_node_id = self.findChild(QLineEdit, 'lineEdit_node_id')
        # QPushButton
        self.pushButton_ChooseFolderImport = self.findChild(QPushButton, 'pushButton_ChooseFolderImport')
        self.pushButton_ChooseFolderExport = self.findChild(QPushButton, 'pushButton_ChooseFolderExport')
        self.pushButton_ExportResults = self.findChild(QPushButton, 'pushButton_ExportResults')
        self.pushButton_ResetPlot = self.findChild(QPushButton, 'pushButton_ResetPlot')
        self.pushButton_AddImportedPlot = self.findChild(QPushButton, 'pushButton_AddImportedPlot')
        self.pushButton_plot = self.findChild(QPushButton, 'pushButton_plot')
        # QRadioButton
        self.radioButton_Absolute = self.findChild(QRadioButton, 'radioButton_Absolute')
        self.radioButton_Real_Imaginary = self.findChild(QRadioButton, 'radioButton_Real_Imaginary')
        self.radioButton_plotAbs = self.findChild(QRadioButton, 'radioButton_plotAbs')
        self.radioButton_plotReal = self.findChild(QRadioButton, 'radioButton_plotReal')
        self.radioButton_plotImag = self.findChild(QRadioButton, 'radioButton_plotImag')
        self.plotAbs = self.radioButton_plotAbs.isChecked()
        self.plotReal = self.radioButton_plotReal.isChecked()
        self.plotImag = self.radioButton_plotImag.isChecked()
        self.save_Absolute = self.radioButton_Absolute.isChecked()
        self.save_Real_Imaginary = self.radioButton_Real_Imaginary.isChecked()
        # QSpinBox
        self.spinBox_skiprows = self.findChild(QSpinBox, 'spinBox')
        # QTabWidget
        self.tabWidget_plot_results = self.findChild(QTabWidget, "tabWidget_plot_results")
        # QWidget
        self.tab_plot = self.tabWidget_plot_results.findChild(QWidget, "tab_plot")


    def _create_connections(self):
        #
        self.pushButton_ChooseFolderImport.clicked.connect(self.choose_path_import_results)
        self.pushButton_ChooseFolderExport.clicked.connect(self.choose_path_export_results)
        self.pushButton_ExportResults.clicked.connect(self.ExportResults)
        self.pushButton_ResetPlot.clicked.connect(self.reset_imported_data)
        self.pushButton_plot.clicked.connect(self.check)
        #
        self.radioButton_plotAbs.clicked.connect(self.radioButtonEvent_YAxis)
        self.radioButton_plotReal.clicked.connect(self.radioButtonEvent_YAxis)
        self.radioButton_plotImag.clicked.connect(self.radioButtonEvent_YAxis)
        self.radioButton_Absolute.clicked.connect(self.radioButtonEvent_save_data)
        self.radioButton_Real_Imaginary.clicked.connect(self.radioButtonEvent_save_data)
        #
        self.pushButton_AddImportedPlot.clicked.connect(self.ImportResults) 


    def reset_imported_data(self):
        self.imported_data = None
        title = "Information"
        message = "The plot data has been reseted."
        PrintMessageInput([title, message, window_title2])
    

    def writeNodes(self, list_node_ids):
        text = ""
        for node in list_node_ids:
            text += "{}, ".format(node)
        self.lineEdit_node_id.setText(text)


    def update(self):
        self.list_node_IDs = self.opv.getListPickedPoints()
        if self.list_node_IDs != []:
            self.writeNodes(self.list_node_IDs)


    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            self.check()
        elif event.key() == Qt.Key_Escape:
            self.close()


    def radioButtonEvent_YAxis(self):
        self.plotAbs = self.radioButton_plotAbs.isChecked()
        self.plotReal = self.radioButton_plotReal.isChecked()
        self.plotImag = self.radioButton_plotImag.isChecked()


    def radioButtonEvent_save_data(self):
        self.save_Absolute = self.radioButton_Absolute.isChecked()
        self.save_Real_Imaginary = self.radioButton_Real_Imaginary.isChecked()


    def choose_path_import_results(self):
        self.import_path, _ = QFileDialog.getOpenFileName(None, 'Open file', self.userPath, 'Files (*.csv; *.dat; *.txt)')
        self.import_name = os.path.basename(self.import_path)
        self.lineEdit_ImportResultsPath.setText(str(self.import_path))


    def import_results(self):
        try:
            message = ""
            run = True
            if self.checkBox_skiprows.isChecked():
                skiprows = self.spinBox_skiprows.value()
            else:
                skiprows = 0
            maximum_lines_to_skip = 100
            
            while run:
                try:
                    sufix = Path(self.imported_path).suffix
                    filename = os.path.basename(self.imported_path)
                    if sufix in [".txt", ".dat", ".csv"]:
                        loaded_data = np.loadtxt(self.imported_path, 
                                                 delimiter = ",", 
                                                 skiprows = skiprows)
                        key = self.get_data_index()
                        self.imported_results[key] = {  "data" : loaded_data,
                                                        "filename" : filename,
                                                        "extension" : sufix  }

                    elif sufix in [".xls", ".xlsx"]:
                        wb = openpyxl.load_workbook(self.imported_path)
                        sheetnames = wb.sheetnames
                        for sheetname in sheetnames:
                            sheet_data = pd.read_excel(self.imported_path, 
                                                       sheet_name = sheetname, 
                                                       header = skiprows, 
                                                       usecols = [0,1,2]).to_numpy()
                            key = self.get_data_index()
                            self.imported_results[key] = {  "data" : sheet_data,
                                                            "filename" : filename,
                                                            "sheetname" : sheetname,
                                                            "extension" : sufix  }

                    self.spinBox_skiprows.setValue(int(skiprows))
                    run = False

                except:
                    skiprows += 1
                    if skiprows >= maximum_lines_to_skip:
                        run = False
                        title = "Error while loading data from file"
                        message = "The maximum number of rows to skip has been reached and no valid data has "
                        message += "been found. Please, verify the data in the imported file to proceed."
                        message += "Maximum number of header rows: 100"

        except Exception as log_error:
            title = "Error while loading data from file"
            message = str(log_error)
            return
        
        if message != "":
            PrintMessageInput([title, message, window_title1])


    def choose_path_export_results(self):
        self.save_path = QFileDialog.getExistingDirectory(None, 'Choose a folder to export the results', self.userPath)
        self.save_name = os.path.basename(self.save_path)
        self.lineEdit_SaveResultsPath.setText(str(self.save_path))


    def check(self, export=False):

        lineEdit_node_id = self.lineEdit_node_id.text()
        stop, self.node_ID = self.before_run.check_input_NodeID(lineEdit_node_id, single_ID=True)
        if stop:
            return True
                
        if self.checkBox_dB.isChecked():
            self.scale_dB = True
        elif not self.checkBox_dB.isChecked():
            self.scale_dB = False
        
        if not export:
            self.plot()

    def ExportResults(self):

        if self.lineEdit_FileName.text() != "":
            if self.save_path != "":
                self.export_path_folder = self.save_path + "/"
            else:
                title = "None folder selected"
                message = "Plese, choose a folder before trying export the results."
                PrintMessageInput([title, message, window_title1])
                return
        else:
            title = "Empty file name"
            message = "Inform a file name before trying export the results."
            PrintMessageInput([title, message, window_title1])
            return

        if self.check(export=True):
            return

        freq = self.frequencies
        self.export_path = self.export_path_folder + self.lineEdit_FileName.text() + ".dat"
        if self.save_Absolute:
            response = get_acoustic_frf(self.preprocessor, self.solution, self.node_ID)
            header = "Frequency[Hz], Real part [Pa], Imaginary part [Pa], Absolute [Pa]"
            data_to_export = np.array([freq, np.real(response), np.imag(response), np.abs(response)]).T
        elif self.save_Real_Imaginary:
            response = get_acoustic_frf(self.preprocessor, self.solution, self.node_ID)
            header = "Frequency[Hz], Real part [Pa], Imaginary part [Pa]"
            data_to_export = np.array([freq, np.real(response), np.imag(response)]).T        
            
        np.savetxt(self.export_path, data_to_export, delimiter=",", header=header)
        title = "Information"
        message = "The results have been exported."
        PrintMessageInput([title, message, window_title2])


    # def dB(self, data):
    #     p_ref = 20e-6 
    #     return 20*np.log10(data/p_ref)


    # def plot(self):
    #     """
    #     """
    #     plt.ion()
    #     # self.fig = plt.figure(figsize=[12,7])
    #     # ax = self.fig.add_subplot(1,1,1)
    #     self.fig, ax = plt.subplots(figsize=(8, 6))

    #     frequencies = self.frequencies
    #     response = get_acoustic_frf(self.preprocessor, self.solution, self.node_ID, absolute=self.plotAbs, real=self.plotReal, imag=self.plotImag)

    #     if complex(0) in response:
    #         response += np.ones(len(response), dtype=float)*(1e-12)

    #     if self.scale_dB :
    #         if self.plotAbs:
    #             response = self.dB(response)
    #             ax.set_ylabel("Acoustic Response - Absolute [dB]", fontsize = 14, fontweight = 'bold')
    #         else:
    #             if self.plotReal:
    #                 ax.set_ylabel("Acoustic Response - Real [Pa]", fontsize = 14, fontweight = 'bold')
    #             elif self.plotImag:
    #                 ax.set_ylabel("Acoustic Response - Imaginary [Pa]", fontsize = 14, fontweight = 'bold')
    #             title = "Plot Information"
    #             message = "The dB scalling can only be applied with the absolute \nY-axis representation, therefore, it will be ignored."
    #             PrintMessageInput([title, message, window_title2])
    #     else:
    #         if self.plotAbs:
    #             ax.set_ylabel("Acoustic Response - Absolute [Pa]", fontsize = 14, fontweight = 'bold')
    #         elif self.plotReal:
    #             ax.set_ylabel("Acoustic Response - Real [Pa]", fontsize = 14, fontweight = 'bold')
    #         elif self.plotImag:
    #             ax.set_ylabel("Acoustic Response - Imaginary [Pa]", fontsize = 14, fontweight = 'bold')

    #     legend_label = "Acoustic Pressure at node {}".format(self.node_ID)
        
    #     if self.imported_data is None:

    #         response += np.ones(len(response), dtype=float)*(1e-12)

    #         if self.plotAbs and not self.scale_dB and not complex(0) in response:
    #             first_plot, = ax.semilogy(frequencies, response, color=[1,0,0], linewidth=2, label=legend_label)
    #         else:
    #             first_plot, = ax.plot(frequencies, response, color=[1,0,0], linewidth=2, label=legend_label)
    #         _legends = ax.legend(handles=[first_plot], labels=[legend_label])

    #     else:

    #         data = self.imported_data
    #         imported_Xvalues = data[:,0]
            
    #         if self.plotAbs:
    #             imported_Yvalues = np.abs(data[:,1] + 1j*data[:,2])
    #             if complex(0) in imported_Yvalues:
    #                 imported_Yvalues += np.ones(len(imported_Yvalues), dtype=float)*(1e-12)

    #             if self.scale_dB :
    #                 imported_Yvalues = self.dB(imported_Yvalues)
    #         elif self.plotReal:
    #             imported_Yvalues = data[:,1]
    #         elif self.plotImag:
    #             imported_Yvalues = data[:,2]

    #         if self.plotAbs and not self.scale_dB and not complex(0) in response:
    #             first_plot, = ax.semilogy(frequencies, response, color=[1,0,0], linewidth=2)
    #             second_plot, = ax.semilogy(imported_Xvalues, imported_Yvalues, color=[0,0,1], linewidth=1, linestyle="--")            
    #         else:
    #             first_plot, = ax.plot(frequencies, response, color=[1,0,0], linewidth=2)
    #             second_plot, = ax.plot(imported_Xvalues, imported_Yvalues, color=[0,0,1], linewidth=1, linestyle="--")

    #         _legends = ax.legend(handles=[first_plot, second_plot], labels=[legend_label, self.legend_imported])#, loc='upper right')

    #     plt.gca().add_artist(_legends)

    #     ax.set_title(('ACOUSTIC FREQUENCY RESPONSE - {}').format(self.analysisMethod.upper()), fontsize = 12, fontweight = 'bold')
    #     ax.set_xlabel(('Frequency [Hz]'), fontsize = 11, fontweight = 'bold')

    #     if not self.radioButton_disable_cursors.isChecked():
    #         show_legend = self.checkBox_legends.isChecked()
    #         if self.radioButton_harmonic_cursor.isChecked():
    #             self.cursor = AdvancedCursor(ax, frequencies, response, False, number_vertLines=12, show_legend=show_legend)
    #         else:
    #             self.cursor = AdvancedCursor(ax, frequencies, response, False, number_vertLines=1 , show_legend=show_legend)

    #         self.mouse_connection = self.fig.canvas.mpl_connect(s='motion_notify_event', func=self.cursor.mouse_move)
        
    #     self.fig.show()


    def call_plotter(self):
        self.join_model_data()
        self.join_imported_data()
        self.plotter = FrequencyResponsePlotter()
        self.plotter._set_data_to_plot(self.data_to_plot)

    def join_model_data(self):
        self.data_to_plot = dict()
        self.title = "Structural frequency response - {}".format(self.analysisMethod)
        legend_label = "Response {} at node {}".format(self.localdof_label, self.node_ID)
        self.data_to_plot[0] = {   "type" : "model results",
                                    "x_data" : self.frequencies,
                                    "y_data" : self.get_response(),
                                    "x_label" : "Frequency [Hz]",
                                    "y_label" : "Nodal response",
                                    "legend" : legend_label,
                                    "unit" : self.unit_label,
                                    "title" : self.title,
                                    "color" : [0,0,1],
                                    "linestyle" : "-"   }

    def join_imported_data(self):
        j = 0
        for id, checkBox in self.ids_to_checkBox.items():
            temp_dict = dict()
            if checkBox.isChecked():

                if id < len(self.colors):
                    color = self.colors[j]
                    j += 1
                else:
                    color = np.random.randint(0,255,3)/255

                data = self.imported_results[id]["data"]
                x_values = data[:, 0]
                y_values = data[:, 1] + 1j*data[:, 2]

                if "sheetname" in self.imported_results[id].keys():
                    sheetname = self.imported_results[id]["sheetname"]
                    legend_label = f"{sheetname}"
                else:
                    legend_label = self.imported_results[id]["filename"]

                temp_dict = {   "type" : "imported_data",
                                "x_data" : x_values,
                                "y_data" : y_values,
                                "x_label" : "Frequency [Hz]",
                                "y_label" : "Nodal response",
                                "legend" : legend_label,
                                "unit" : self.unit_label,
                                "title" : self.title,
                                "color" : color,
                                "linestyle" : "--"   }

                self.data_to_plot[id] = temp_dict